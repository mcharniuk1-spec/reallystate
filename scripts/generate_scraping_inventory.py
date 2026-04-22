#!/usr/bin/env python3
"""Generate scraping inventory: XLSX table + MD summary + PDF report.

Reads source_registry.json and scans tests/fixtures/ to produce per-source
stats on categories, fixture counts, photos, text fields, geo, etc.

Outputs:
  docs/exports/scraping-inventory.xlsx
  docs/exports/scraping-inventory-summary.md
  docs/exports/scraping-inventory-summary.pdf
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
REGISTRY_PATH = REPO / "data" / "source_registry.json"
FIXTURES_ROOT = REPO / "tests" / "fixtures"
EXPORTS = REPO / "docs" / "exports"

FIXTURE_DIR_TO_SOURCE: dict[str, str] = {
    "homes_bg": "Homes.bg",
    "olx_bg": "OLX.bg",
    "imot_bg": "imot.bg",
    "imoti_net": "imoti.net",
    "alo_bg": "alo.bg",
    "address_bg": "Address.bg",
    "bulgarianproperties": "BulgarianProperties",
    "suprimmo": "SUPRIMMO",
    "luximmo": "LUXIMMO",
    "property_bg": "property.bg",
    "bazar_bg": "Bazar.bg",
    "domaza": "Domaza",
    "yavlena": "Yavlena",
    "home2u": "Home2U",
    "bcpea": "BCPEA property auctions",
    "telegram_public": "Telegram public channels",
    "x_public": "X public search/accounts",
    "tier3_templates": "(Tier-3 templates)",
}

CATEGORY_NICE = {
    "sale": "Sale",
    "long_term_rent": "Long-term Rent",
    "short_term_rent": "Short-term Rent",
    "land": "Land",
    "new_build": "New Build",
    "auction_sale": "Auction Sale",
    "leads": "Leads / Intelligence",
    "short_term_rent_metrics": "STR Metrics",
    "ownership_verification": "Ownership Verification",
    "building_geometry": "Building Geometry",
    "parcel_validation": "Parcel Validation",
    "branding": "Branding",
    "news": "News",
    "links": "Links",
    "property_management": "Property Management",
}

# Estimated total listing counts per source (public data / industry estimates)
ESTIMATED_TOTAL_LISTINGS: dict[str, int | str] = {
    "Homes.bg": 120_000,
    "imot.bg": 200_000,
    "OLX.bg": 45_000,
    "alo.bg": 35_000,
    "imoti.net": 90_000,
    "Address.bg": 18_000,
    "BulgarianProperties": 12_000,
    "SUPRIMMO": 15_000,
    "LUXIMMO": 4_000,
    "property.bg": 10_000,
    "Bazar.bg": 8_000,
    "Domaza": 22_000,
    "Yavlena": 9_000,
    "Home2U": 6_000,
    "Holding Group Real Estate": 3_000,
    "Imoteka.bg": 14_000,
    "Imoti.info": 25_000,
    "Indomio.bg": 7_000,
    "Lions Group": 2_500,
    "Pochivka.bg": 5_000,
    "realestates.bg": 6_000,
    "Realistimo": 4_500,
    "Rentica.bg": 1_200,
    "Svobodni-kvartiri.com": 3_500,
    "Unique Estates": 2_000,
    "Vila.bg": 4_000,
    "ApartmentsBulgaria.com": 1_800,
    "Airbnb": "N/A (partner)",
    "Booking.com": "N/A (partner)",
    "Vrbo": "N/A (partner)",
    "Flat Manager": "N/A (partner)",
    "Menada Bulgaria": "N/A (partner)",
    "AirDNA": "N/A (licensed)",
    "Airbtics": "N/A (licensed)",
    "BCPEA property auctions": 1_224,
    "Property Register": "N/A (manual)",
    "KAIS Cadastre": "N/A (manual)",
    "Facebook public groups/pages": "N/A (social)",
    "Instagram public profiles": "N/A (social)",
    "Telegram public channels": "N/A (social)",
    "Threads public profiles": "N/A (social)",
    "Viber opt-in communities": "N/A (social)",
    "WhatsApp opt-in groups": "N/A (social)",
    "X public search/accounts": "N/A (social)",
}


@dataclass
class FixtureStats:
    listing_cases: int = 0
    discovery_cases: int = 0
    blocked_cases: int = 0
    total_photos: int = 0
    has_description: int = 0
    has_geo: int = 0
    has_address: int = 0
    has_area: int = 0
    has_rooms: int = 0
    has_price: int = 0
    intents_seen: set[str] = field(default_factory=set)
    categories_seen: set[str] = field(default_factory=set)
    discovery_urls_total: int = 0


@dataclass
class LiveScrapeStats:
    urls_discovered: int = 0
    listings_parsed: int = 0
    photos_downloaded: int = 0
    last_run_utc: str = ""
    product_breakdown: dict[str, int] = field(default_factory=dict)


def _read_json(p: Path) -> dict:
    with open(p) as f:
        return json.load(f)


def scan_fixtures() -> dict[str, FixtureStats]:
    stats: dict[str, FixtureStats] = {}
    if not FIXTURES_ROOT.exists():
        return stats

    for src_dir in sorted(FIXTURES_ROOT.iterdir()):
        if not src_dir.is_dir():
            continue
        source_name = FIXTURE_DIR_TO_SOURCE.get(src_dir.name, src_dir.name)
        st = FixtureStats()
        for case_dir in sorted(src_dir.iterdir()):
            if not case_dir.is_dir():
                continue
            exp_path = case_dir / "expected.json"
            if not exp_path.exists():
                continue
            data = _read_json(exp_path)
            case = case_dir.name.lower()

            if "discovery" in case:
                st.discovery_cases += 1
                urls = data.get("urls") or data.get("entries") or data.get("listings") or []
                st.discovery_urls_total += len(urls)
            elif "blocked" in case or "removed" in case or "404" in case:
                st.blocked_cases += 1
            else:
                st.listing_cases += 1
                photos = data.get("image_urls") or []
                st.total_photos += len(photos)
                if data.get("description_snippet") or data.get("description_text"):
                    st.has_description += 1
                lat = data.get("latitude")
                lon = data.get("longitude")
                if lat is not None and lon is not None:
                    st.has_geo += 1
                if data.get("address_text") or data.get("city"):
                    st.has_address += 1
                if data.get("area_sqm") is not None:
                    st.has_area += 1
                if data.get("rooms") is not None:
                    st.has_rooms += 1
                price = data.get("price") or data.get("price_amount")
                if price is not None:
                    st.has_price += 1
                intent = data.get("listing_intent") or data.get("intent")
                if intent:
                    st.intents_seen.add(intent)
                cat = data.get("property_category") or data.get("property_type")
                if cat:
                    st.categories_seen.add(cat)

        stats[source_name] = st
    return stats


def scan_live_scrapes() -> dict[str, LiveScrapeStats]:
    root = REPO / "data" / "scraped"
    stats: dict[str, LiveScrapeStats] = {}
    if not root.exists():
        return stats

    for source_dir in sorted(root.iterdir()):
        if not source_dir.is_dir():
            continue
        stats_path = source_dir / "scrape_stats.json"
        if not stats_path.exists():
            continue
        raw = _read_json(stats_path)
        source_name = raw.get("source_name") or source_dir.name
        stats[source_name] = LiveScrapeStats(
            urls_discovered=int(raw.get("listing_urls_discovered", 0) or 0),
            listings_parsed=int(raw.get("listing_pages_parsed", 0) or 0),
            photos_downloaded=int(raw.get("photos_downloaded", 0) or 0),
            last_run_utc=str(raw.get("end_time") or raw.get("start_time") or ""),
            product_breakdown=dict(raw.get("product_breakdown") or {}),
        )
    return stats


def load_registry() -> list[dict]:
    return _read_json(REGISTRY_PATH)["sources"]


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

def generate_xlsx(
    sources: list[dict],
    fstats: dict[str, FixtureStats],
    lstats: dict[str, LiveScrapeStats],
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Scraping Inventory"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    tier_fills = {
        1: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        2: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        3: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        4: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }

    headers = [
        "Source Name",
        "Tier",
        "Family",
        "URL",
        "Access Mode",
        "Legal Mode",
        "Risk",
        "Categories to Scrape",
        "Est. Total Listings",
        "Live URLs Discovered",
        "Live Listings Parsed",
        "Live Photos Downloaded",
        "Latest Live Run (UTC)",
        "Live Product Buckets",
        "Fixture Listings Scraped",
        "Discovery Fixtures",
        "Discovery URLs Found",
        "Blocked/Error Fixtures",
        "Photos Scraped",
        "Has Description",
        "Has Geo (lat/lon)",
        "Has Address/City",
        "Has Area (sqm)",
        "Has Rooms",
        "Has Price",
        "Intents Covered",
        "Property Types Seen",
        "Connector Status",
        "Notes",
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    connector_status_map = {
        "Homes.bg": "Full connector + discovery",
        "OLX.bg": "Full connector (API) + discovery",
        "imot.bg": "Scaffold + discovery",
        "imoti.net": "Scaffold (headless needed)",
        "alo.bg": "Scaffold + discovery",
        "Address.bg": "Scaffold + discovery",
        "BulgarianProperties": "Scaffold + discovery",
        "SUPRIMMO": "Scaffold + discovery",
        "LUXIMMO": "Scaffold + discovery",
        "property.bg": "Scaffold + discovery",
        "Bazar.bg": "Tier-2 stub connector",
        "Domaza": "Tier-2 stub connector",
        "Yavlena": "Tier-2 stub connector",
        "Home2U": "Tier-2 stub connector",
        "BCPEA property auctions": "Tier-3 stub (auction parser)",
    }

    for row_idx, src in enumerate(sources, 2):
        name = src["source_name"]
        tier = src["tier"]
        fs = fstats.get(name, FixtureStats())
        ls = lstats.get(name, LiveScrapeStats())
        est = ESTIMATED_TOTAL_LISTINGS.get(name, "Unknown")
        cats = [CATEGORY_NICE.get(c, c) for c in src.get("listing_types", [])]
        status = connector_status_map.get(name, "Not yet implemented" if tier <= 2 else "Blocked / deferred")
        if ls.listings_parsed:
            status += " + live batch"

        row = [
            name,
            tier,
            src.get("source_family", ""),
            src.get("primary_url", ""),
            src.get("access_mode", ""),
            src.get("legal_mode", ""),
            src.get("risk_mode", ""),
            ", ".join(cats),
            est if isinstance(est, str) else est,
            ls.urls_discovered,
            ls.listings_parsed,
            ls.photos_downloaded,
            ls.last_run_utc,
            ", ".join(f"{k}={v}" for k, v in sorted(ls.product_breakdown.items())) if ls.product_breakdown else "-",
            fs.listing_cases,
            fs.discovery_cases,
            fs.discovery_urls_total,
            fs.blocked_cases,
            fs.total_photos,
            fs.has_description,
            fs.has_geo,
            fs.has_address,
            fs.has_area,
            fs.has_rooms,
            fs.has_price,
            ", ".join(sorted(fs.intents_seen)) if fs.intents_seen else "-",
            ", ".join(sorted(fs.categories_seen)) if fs.categories_seen else "-",
            status,
            src.get("notes", ""),
        ]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if tier in tier_fills:
                cell.fill = tier_fills[tier]

    col_widths = [22, 5, 12, 32, 14, 26, 10, 36, 16, 12, 12, 12, 18, 28, 12, 10, 12, 10, 8, 8, 10, 12, 10, 8, 8, 22, 22, 26, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sources) + 1}"
    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    summary_rows = [
        ("Report Generated", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Total Sources", len(sources)),
        ("Tier-1 Sources", sum(1 for s in sources if s["tier"] == 1)),
        ("Tier-2 Sources", sum(1 for s in sources if s["tier"] == 2)),
        ("Tier-3 Sources", sum(1 for s in sources if s["tier"] == 3)),
        ("Tier-4 Sources (Social)", sum(1 for s in sources if s["tier"] == 4)),
        ("", ""),
        ("Total Fixture Listing Cases", sum(f.listing_cases for f in fstats.values())),
        ("Total Live Listings Parsed", sum(f.listings_parsed for f in lstats.values())),
        ("Total Live URLs Discovered", sum(f.urls_discovered for f in lstats.values())),
        ("Total Live Photos Downloaded", sum(f.photos_downloaded for f in lstats.values())),
        ("Total Discovery Fixtures", sum(f.discovery_cases for f in fstats.values())),
        ("Total Discovery URLs Extracted", sum(f.discovery_urls_total for f in fstats.values())),
        ("Total Blocked/Error Fixtures", sum(f.blocked_cases for f in fstats.values())),
        ("Total Photos in Fixtures", sum(f.total_photos for f in fstats.values())),
        ("Fixtures with Geo", sum(f.has_geo for f in fstats.values())),
        ("Fixtures with Address", sum(f.has_address for f in fstats.values())),
        ("Fixtures with Price", sum(f.has_price for f in fstats.values())),
        ("Fixtures with Area", sum(f.has_area for f in fstats.values())),
    ]
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20
    for r, (label, val) in enumerate(summary_rows, 1):
        lc = ws2.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True) if label else Font()
        ws2.cell(row=r, column=2, value=val)

    out = EXPORTS / "scraping-inventory.xlsx"
    wb.save(str(out))
    return out


# ---------------------------------------------------------------------------
# Markdown generation
# ---------------------------------------------------------------------------

def generate_md(
    sources: list[dict],
    fstats: dict[str, FixtureStats],
    lstats: dict[str, LiveScrapeStats],
) -> Path:
    lines: list[str] = []
    now = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines.append("# Bulgaria Real Estate - Scraping Inventory & Source Summary\n")
    lines.append(f"*Generated: {now}*\n")

    # Global summary
    total_listing = sum(f.listing_cases for f in fstats.values())
    total_disc = sum(f.discovery_cases for f in fstats.values())
    total_photos = sum(f.total_photos for f in fstats.values())
    total_live_parsed = sum(f.listings_parsed for f in lstats.values())
    total_live_discovered = sum(f.urls_discovered for f in lstats.values())
    lines.append("## Overview\n")
    lines.append("| Metric | Value |")
    lines.append("|--------|-------|")
    lines.append(f"| Total sources in registry | {len(sources)} |")
    lines.append(f"| Tier-1 (primary portals) | {sum(1 for s in sources if s['tier']==1)} |")
    lines.append(f"| Tier-2 (expansion) | {sum(1 for s in sources if s['tier']==2)} |")
    lines.append(f"| Tier-3 (partner/official) | {sum(1 for s in sources if s['tier']==3)} |")
    lines.append(f"| Tier-4 (social/lead intel) | {sum(1 for s in sources if s['tier']==4)} |")
    lines.append(f"| Total listing fixtures scraped | {total_listing} |")
    lines.append(f"| Total live listings parsed | {total_live_parsed} |")
    lines.append(f"| Total live URLs discovered | {total_live_discovered} |")
    lines.append(f"| Total discovery fixtures | {total_disc} |")
    lines.append(f"| Total photos captured in fixtures | {total_photos} |")
    lines.append("")

    for tier in [1, 2, 3, 4]:
        tier_sources = [s for s in sources if s["tier"] == tier]
        if not tier_sources:
            continue
        tier_label = {1: "Tier 1 - Primary Portals & Classifieds", 2: "Tier 2 - Expansion Sources",
                      3: "Tier 3 - Partner / Official / Vendor", 4: "Tier 4 - Social & Lead Intelligence"}
        lines.append(f"## {tier_label.get(tier, f'Tier {tier}')}\n")

        for src in tier_sources:
            name = src["source_name"]
            url = src.get("primary_url", "")
            fs = fstats.get(name, FixtureStats())
            ls = lstats.get(name, LiveScrapeStats())
            cats = [CATEGORY_NICE.get(c, c) for c in src.get("listing_types", [])]
            est = ESTIMATED_TOTAL_LISTINGS.get(name, "Unknown")

            lines.append(f"### [{name}]({url})\n")
            lines.append(f"- **URL**: {url}")
            lines.append(f"- **Family**: {src.get('source_family', '-')}")
            lines.append(f"- **Owner**: {src.get('owner_group', '-')}")
            lines.append(f"- **Access**: `{src.get('access_mode', '-')}`")
            lines.append(f"- **Legal mode**: `{src.get('legal_mode', '-')}`")
            lines.append(f"- **Risk**: `{src.get('risk_mode', '-')}`")
            lines.append(f"- **Languages**: {', '.join(src.get('languages', []))}")
            lines.append(f"- **Categories to scrape**: {', '.join(cats)}")
            lines.append(f"- **Est. total listings on site**: {est}")
            lines.append(f"- **Extraction method**: {src.get('best_extraction_method', '-')}")
            lines.append(f"- **Notes**: {src.get('notes', '-')}")

            if ls.listings_parsed or ls.urls_discovered:
                lines.append("\n**Live Harvest Statistics:**\n")
                lines.append("| Metric | Count |")
                lines.append("|--------|-------|")
                lines.append(f"| Live URLs discovered | {ls.urls_discovered} |")
                lines.append(f"| Live listings parsed | {ls.listings_parsed} |")
                lines.append(f"| Live photos downloaded | {ls.photos_downloaded} |")
                lines.append(f"| Latest live run | {ls.last_run_utc or '-'} |")
                lines.append(
                    f"| Live product buckets | {', '.join(f'{k}={v}' for k, v in sorted(ls.product_breakdown.items())) or '-'} |"
                )

            if fs.listing_cases or fs.discovery_cases or fs.blocked_cases:
                lines.append("\n**Fixture Statistics:**\n")
                lines.append("| Metric | Count |")
                lines.append("|--------|-------|")
                lines.append(f"| Listing fixtures | {fs.listing_cases} |")
                lines.append(f"| Discovery fixtures | {fs.discovery_cases} |")
                lines.append(f"| Discovery URLs extracted | {fs.discovery_urls_total} |")
                lines.append(f"| Blocked/error fixtures | {fs.blocked_cases} |")
                lines.append(f"| Photos captured | {fs.total_photos} |")
                lines.append(f"| With geo (lat/lon) | {fs.has_geo} |")
                lines.append(f"| With address/city | {fs.has_address} |")
                lines.append(f"| With price | {fs.has_price} |")
                lines.append(f"| With area (sqm) | {fs.has_area} |")
                lines.append(f"| With rooms | {fs.has_rooms} |")
                lines.append(f"| Intents covered | {', '.join(sorted(fs.intents_seen)) or '-'} |")
                lines.append(f"| Property types seen | {', '.join(sorted(fs.categories_seen)) or '-'} |")
            else:
                lines.append("\n*No fixtures yet.*\n")

            lines.append("")

    lines.append("---\n")
    lines.append("## Legend\n")
    lines.append("- **Tier 1**: Highest-volume national portals and classifieds. First to be fully connected.")
    lines.append("- **Tier 2**: Regional agencies, secondary portals, travel/OTA sites. Expanded after tier-1 is stable.")
    lines.append("- **Tier 3**: Partner feeds (Airbnb, Booking.com), licensed data vendors (AirDNA), official registers (BCPEA, KAIS). Require contracts or manual consent.")
    lines.append("- **Tier 4**: Social and messaging platforms. Lead intelligence overlay only; consent-gated.")
    lines.append("- **Fixture**: Offline HTML/JSON snapshot used for parser testing without live network calls.")
    lines.append("- **Discovery fixture**: Simulates a search results page to test URL extraction and pagination.")
    lines.append("- **Blocked fixture**: Simulates anti-bot / access-denied responses for error handling tests.")
    lines.append("")

    out = EXPORTS / "scraping-inventory-summary.md"
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

def generate_pdf(
    sources: list[dict],
    fstats: dict[str, FixtureStats],
    lstats: dict[str, LiveScrapeStats],
) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = EXPORTS / "scraping-inventory-summary.pdf"
    doc = SimpleDocTemplate(
        str(out),
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18, spaceAfter=12)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, spaceAfter=6, spaceBefore=12)
    body_style = ParagraphStyle("Body2", parent=styles["Normal"], fontSize=8, leading=10)
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7, leading=9)
    link_style = ParagraphStyle("Link", parent=small, textColor=colors.HexColor("#2F5496"))

    elements: list = []
    now = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    elements.append(Paragraph("Bulgaria Real Estate - Scraping Inventory", title_style))
    elements.append(Paragraph(f"Generated: {now}", body_style))
    elements.append(Spacer(1, 8 * mm))

    # Summary table
    total_listing = sum(f.listing_cases for f in fstats.values())
    total_disc = sum(f.discovery_cases for f in fstats.values())
    total_photos = sum(f.total_photos for f in fstats.values())
    total_live_parsed = sum(f.listings_parsed for f in lstats.values())

    sum_data = [
        ["Metric", "Value"],
        ["Total sources", str(len(sources))],
        ["Tier-1", str(sum(1 for s in sources if s["tier"] == 1))],
        ["Tier-2", str(sum(1 for s in sources if s["tier"] == 2))],
        ["Tier-3", str(sum(1 for s in sources if s["tier"] == 3))],
        ["Tier-4", str(sum(1 for s in sources if s["tier"] == 4))],
        ["Listing fixtures", str(total_listing)],
        ["Live listings parsed", str(total_live_parsed)],
        ["Discovery fixtures", str(total_disc)],
        ["Photos in fixtures", str(total_photos)],
    ]
    st = Table(sum_data, colWidths=[5 * cm, 3 * cm])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 6 * mm))

    tier_colors = {
        1: colors.HexColor("#E2EFDA"),
        2: colors.HexColor("#FFF2CC"),
        3: colors.HexColor("#FCE4D6"),
        4: colors.HexColor("#F2F2F2"),
    }

    for tier in [1, 2, 3, 4]:
        tier_sources = [s for s in sources if s["tier"] == tier]
        if not tier_sources:
            continue
        tier_label = {1: "Tier 1 - Primary Portals", 2: "Tier 2 - Expansion",
                      3: "Tier 3 - Partner/Official", 4: "Tier 4 - Social/Leads"}
        elements.append(Paragraph(tier_label.get(tier, f"Tier {tier}"), h2_style))

        table_data = [
            [
                Paragraph("<b>Source</b>", small),
                Paragraph("<b>URL</b>", small),
                Paragraph("<b>Categories</b>", small),
                Paragraph("<b>Est. Total</b>", small),
                Paragraph("<b>Listings Scraped</b>", small),
                Paragraph("<b>Discovery</b>", small),
                Paragraph("<b>Photos</b>", small),
                Paragraph("<b>Geo</b>", small),
                Paragraph("<b>Price</b>", small),
                Paragraph("<b>Area</b>", small),
                Paragraph("<b>Access</b>", small),
                Paragraph("<b>Legal</b>", small),
            ],
        ]

        for src in tier_sources:
            name = src["source_name"]
            url = src.get("primary_url", "")
            fs = fstats.get(name, FixtureStats())
            cats = [CATEGORY_NICE.get(c, c) for c in src.get("listing_types", [])]
            est = ESTIMATED_TOTAL_LISTINGS.get(name, "?")
            url_text = url[:40] + "..." if len(url) > 43 else url

            table_data.append([
                Paragraph(name, small),
                Paragraph(f'<a href="{url}" color="#2F5496">{url_text}</a>', link_style),
                Paragraph("<br/>".join(cats), small),
                Paragraph(str(est), small),
                Paragraph(str(fs.listing_cases), small),
                Paragraph(str(fs.discovery_cases), small),
                Paragraph(str(fs.total_photos), small),
                Paragraph(str(fs.has_geo), small),
                Paragraph(str(fs.has_price), small),
                Paragraph(str(fs.has_area), small),
                Paragraph(src.get("access_mode", ""), small),
                Paragraph(src.get("legal_mode", ""), small),
            ])

        col_w = [2.6 * cm, 3.8 * cm, 3.2 * cm, 1.8 * cm, 1.5 * cm, 1.3 * cm, 1.1 * cm, 0.9 * cm, 0.9 * cm, 0.9 * cm, 2.2 * cm, 3.2 * cm]
        t = Table(table_data, colWidths=col_w, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        tc = tier_colors.get(tier)
        if tc:
            style_cmds.append(("BACKGROUND", (0, 1), (-1, -1), tc))
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
        elements.append(Spacer(1, 4 * mm))

    # Legend
    elements.append(Paragraph("Legend", h2_style))
    legend_items = [
        "<b>Tier 1</b>: Highest-volume national portals. First fully connected.",
        "<b>Tier 2</b>: Regional agencies, secondary portals, OTA sites.",
        "<b>Tier 3</b>: Partner feeds, licensed vendors, official registers. Require contracts.",
        "<b>Tier 4</b>: Social/messaging. Lead intelligence overlay; consent-gated.",
        "<b>Listings Scraped</b>: Offline fixture cases parsed and validated.",
        "<b>Discovery</b>: Search-page parsing fixtures for URL extraction.",
        "<b>Photos</b>: Count of image URLs captured in listing fixtures.",
        "<b>Geo</b>: Fixtures with latitude/longitude coordinates.",
    ]
    for item in legend_items:
        elements.append(Paragraph(item, body_style))
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        "This report is auto-generated from source_registry.json and tests/fixtures/. "
        "Estimated total listings are industry approximations and may vary.",
        ParagraphStyle("Disclaimer", parent=body_style, textColor=colors.grey, fontSize=7),
    ))

    doc.build(elements)
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    sources = load_registry()
    fstats = scan_fixtures()
    lstats = scan_live_scrapes()

    xlsx_path = generate_xlsx(sources, fstats, lstats)
    print(f"XLSX: {xlsx_path}")

    md_path = generate_md(sources, fstats, lstats)
    print(f"MD:   {md_path}")

    pdf_path = generate_pdf(sources, fstats, lstats)
    print(f"PDF:  {pdf_path}")

    print(f"\nDone. {len(sources)} sources, {sum(f.listing_cases for f in fstats.values())} listing fixtures, "
          f"{sum(f.listings_parsed for f in lstats.values())} live listings parsed, "
          f"{sum(f.discovery_cases for f in fstats.values())} discovery fixtures, "
          f"{sum(f.total_photos for f in fstats.values())} photos.")


if __name__ == "__main__":
    main()
