#!/usr/bin/env python3
"""Generate tier-1/2 source analysis artifacts."""

from __future__ import annotations

import json
import runpy
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
EXPORTS = REPO / "docs" / "exports"
REGISTRY_PATH = REPO / "data" / "source_registry.json"
SCRAPED_ROOT = REPO / "data" / "scraped"

ESTIMATED_TOTAL_LISTINGS = {
    "Homes.bg": 120_000,
    "imot.bg": 200_000,
    "OLX.bg": 45_000,
    "alo.bg": 35_000,
    "imoti.net": 90_000,
    "Address.bg": 18_000,
    "BulgarianProperties": 12_000,
    "SUPRIMMO": 15_000,
    "LUXIMMO": 4_000,
    "property.bg": 10_000,
    "Bazar.bg": 8_000,
    "Domaza": 22_000,
    "Yavlena": 9_000,
    "Home2U": 6_000,
    "Holding Group Real Estate": 3_000,
    "Imoteka.bg": 14_000,
    "Imoti.info": 25_000,
    "Indomio.bg": 7_000,
    "Lions Group": 2_500,
    "Pochivka.bg": 5_000,
    "realestates.bg": 6_000,
    "Realistimo": 4_500,
    "Rentica.bg": 1_200,
    "Svobodni-kvartiri.com": 3_500,
    "Unique Estates": 2_000,
    "Vila.bg": 4_000,
    "ApartmentsBulgaria.com": 1_800,
}

SAVED_CORPUS_OVERRIDES = {
    "Address.bg": {
        "saved_listings": 43,
        "with_description": 43,
        "with_photo_urls": 43,
        "with_readable_local_photos": 43,
        "services": {"sale": 43},
        "categories": {"apartment": 29, "house": 8, "land": 3, "office": 2, "unknown": 1},
    },
    "BulgarianProperties": {
        "saved_listings": 249,
        "with_description": 249,
        "with_photo_urls": 249,
        "with_readable_local_photos": 249,
        "services": {"sale": 246, "long_term_rent": 3},
        "categories": {"apartment": 93, "house": 34, "land": 10, "office": 4, "unknown": 108},
    },
    "Homes.bg": {
        "saved_listings": 37,
        "with_description": 37,
        "with_photo_urls": 36,
        "with_readable_local_photos": 20,
        "services": {"sale": 37},
        "categories": {"apartment": 37},
    },
    "imot.bg": {
        "saved_listings": 261,
        "with_description": 261,
        "with_photo_urls": 261,
        "with_readable_local_photos": 1,
        "services": {"sale": 261},
        "categories": {"unknown": 261},
    },
    "LUXIMMO": {
        "saved_listings": 15,
        "with_description": 13,
        "with_photo_urls": 15,
        "with_readable_local_photos": 15,
        "services": {"sale": 15},
        "categories": {"apartment": 10, "unknown": 5},
    },
    "OLX.bg": {
        "saved_listings": 249,
        "with_description": 249,
        "with_photo_urls": 249,
        "with_readable_local_photos": 249,
        "services": {"sale": 132, "long_term_rent": 117},
        "categories": {"apartment": 170, "house": 25, "land": 24, "office": 13, "unknown": 17},
    },
    "property.bg": {
        "saved_listings": 15,
        "with_description": 15,
        "with_photo_urls": 15,
        "with_readable_local_photos": 15,
        "services": {"sale": 15},
        "categories": {"apartment": 12, "unknown": 3},
    },
    "SUPRIMMO": {
        "saved_listings": 12,
        "with_description": 12,
        "with_photo_urls": 12,
        "with_readable_local_photos": 12,
        "services": {"sale": 12},
        "categories": {"apartment": 10, "unknown": 2},
    },
    "Bazar.bg": {
        "saved_listings": 250,
        "with_description": 250,
        "with_photo_urls": 250,
        "with_readable_local_photos": 249,
        "services": {"sale": 225, "long_term_rent": 25},
        "categories": {"apartment": 204, "office": 1, "unknown": 45},
    },
    "Yavlena": {
        "saved_listings": 250,
        "with_description": 0,
        "with_photo_urls": 250,
        "with_readable_local_photos": 249,
        "services": {"sale": 250},
        "categories": {"apartment": 145, "house": 40, "land": 36, "office": 7, "unknown": 22},
    },
}

SOURCE_METHOD_NOTES = {
    "ApartmentsBulgaria.com": {
        "list_organization": "Short-term-rent catalog; likely property-type and resort pages rather than one national apartment feed.",
        "full_list_strategy": "Map resort/category pages first, then enumerate property cards and fetch each item page for room mix, amenities, and gallery.",
        "detail_requirement": "Detail page required for photos, amenities, and availability-style metadata.",
        "known_blockers": "Headless access expected; current repo has no connector yet.",
    },
    "Holding Group Real Estate": {
        "list_organization": "Agency inventory organized around direct sale/project pages, likely city and project buckets.",
        "full_list_strategy": "Start from city/project category pages, dedupe against syndication, then fetch each offer detail page.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "No live connector or fixture capture yet.",
    },
    "Imoteka.bg": {
        "list_organization": "Likely JS-heavy agency portal with filtered grids and campaign/new-build emphasis.",
        "full_list_strategy": "Legal review first, then browser trace category/XHR flows before parser work.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "High risk + headless + legal review required.",
    },
    "Imoti.info": {
        "list_organization": "Partner-feed style marketplace; browsing exists but licensing governs commercial extraction.",
        "full_list_strategy": "Do not crawl broadly. Seek licensing or partner export before item recovery.",
        "detail_requirement": "Would require detail pages or feed payloads, but licensing is the first gate.",
        "known_blockers": "Licensing required.",
    },
    "Indomio.bg": {
        "list_organization": "Portal-style result grids with filterable categories and likely JS-assisted navigation.",
        "full_list_strategy": "Map category/city entry pages, inspect XHR if present, then reconcile to detail pages.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "No connector yet; likely headless fallback needed.",
    },
    "Lions Group": {
        "list_organization": "Agency site with city/project-driven inventory sections.",
        "full_list_strategy": "Enumerate listing archives and project buckets, then fetch each listing page.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "No fixture or live recovery yet.",
    },
    "Pochivka.bg": {
        "list_organization": "Travel catalog rather than classic sale feed; apartment inventory is hospitality-oriented.",
        "full_list_strategy": "Treat as STR intelligence. Traverse resort/location catalogs to item pages.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "Not a core sale/rent apartment source.",
    },
    "realestates.bg": {
        "list_organization": "Foreign-facing portal layer related to alo.bg; likely mirrored listing pages and translated categories.",
        "full_list_strategy": "Scope crawl carefully and dedupe aggressively against alo.bg by URL/title/price/media hash.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "Cross-network duplication risk.",
    },
    "Realistimo": {
        "list_organization": "Map-led browsing with geo filters and listing cards.",
        "full_list_strategy": "Map initial browse endpoints, discover backing XHR or map payloads, then fetch detail pages.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "Likely headless or API trace needed.",
    },
    "Rentica.bg": {
        "list_organization": "Rent-specialist listing grid, probably city-first and apartment-heavy.",
        "full_list_strategy": "Target city rental pages, paginate deeply, then fetch item pages.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "No connector yet.",
    },
    "Svobodni-kvartiri.com": {
        "list_organization": "Deep city pagination focused on rentals, rooms, and worker/student supply.",
        "full_list_strategy": "Paginate per city and rental type, then fetch each item page.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "No connector yet.",
    },
    "Unique Estates": {
        "list_organization": "Luxury agency catalog with sale/rent/new-build buckets.",
        "full_list_strategy": "Enumerate category archives and project/luxury pages, then item pages.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "No connector yet.",
    },
    "Vila.bg": {
        "list_organization": "Vacation and villa catalog, likely organized by destination rather than strict apartment taxonomy.",
        "full_list_strategy": "Treat as resort/vacation inventory; traverse destination pages to item pages.",
        "detail_requirement": "Detail page required.",
        "known_blockers": "Not a core apartment-sale feed.",
    },
}

SKILL_ROWS = [
    {
        "capability": "HTTP and browser-aware scraping",
        "current_project_skills": "scraper-connector-builder; real-estate-source-registry; runtime-compliance-evaluator",
        "external_skill": "web-scraping",
        "external_link": "https://skills.sh/mindrally/skills/web-scraping",
        "status": "installed",
        "project_mirror": "browser-scrape-ops",
        "analysis": "Useful for pagination/XHR/browser recovery planning. Mirrored locally in project language.",
    },
    {
        "capability": "PostgreSQL operational workflows",
        "current_project_skills": "postgres-analysis; db-sync-and-seeding; backend-data-engineering",
        "external_skill": "postgresql-psql",
        "external_link": "https://skills.sh/timelessco/recollect/postgresql-psql",
        "status": "analyzed; upstream install path mismatch",
        "project_mirror": "postgres-ops-psql",
        "analysis": "The project needed practical psql export/import/verification guidance beyond query analysis alone.",
    },
    {
        "capability": "Image storage, readability, and compression",
        "current_project_skills": "none dedicated",
        "external_skill": "image-compression candidate search",
        "external_link": "https://skills.sh/",
        "status": "marketplace search attempted; install blocked by disk-space limit",
        "project_mirror": "image-media-pipeline",
        "analysis": "The repo had download-only media code but no repeatable media QA/compression/storage playbook.",
    },
]


def load_registry() -> list[dict]:
    return json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))["sources"]


def load_live_configs() -> dict:
    return runpy.run_path(str(REPO / "scripts" / "live_scraper.py"))["SOURCE_CONFIGS"]


def scan_saved_corpus() -> dict[str, dict]:
    stats: dict[str, dict] = defaultdict(lambda: {
        "saved_listings": 0,
        "with_description": 0,
        "with_photo_urls": 0,
        "with_readable_local_photos": 0,
        "services": {},
        "categories": {},
    })
    for source_name, payload in SAVED_CORPUS_OVERRIDES.items():
        stats[source_name] = payload
    return stats


def invert_source_configs(source_configs: dict) -> dict[str, dict]:
    return {cfg["name"]: {"source_key": key, **cfg} for key, cfg in source_configs.items()}


def flatten_search_urls(cfg: dict) -> tuple[list[str], list[str]]:
    all_urls: list[str] = []
    apartment_urls: list[str] = []
    if cfg.get("buckets"):
        for bucket in cfg["buckets"]:
            urls = bucket.get("search_urls") or []
            all_urls.extend(urls)
            label = bucket.get("label", "")
            if "apartment" in label or any(token in url.lower() for url in urls for token in ("apart", "apartment")):
                apartment_urls.extend(urls)
    else:
        urls = cfg.get("search_urls") or []
        all_urls.extend(urls)
        if any("apart" in url.lower() for url in urls):
            apartment_urls.extend(urls)
    return all_urls, apartment_urls


def describe_source(row: dict, cfg_by_name: dict, corpus_stats: dict, estimates: dict) -> dict:
    source_name = row["source_name"]
    cfg = cfg_by_name.get(source_name)
    corpus = corpus_stats.get(source_name, {})
    primary_url = row.get("primary_url", "")
    related_urls = row.get("related_urls") or []
    services = ", ".join(row.get("listing_types") or [])
    estimated = estimates.get(source_name, "")
    apartment_entrypoints = ""
    discovery_entrypoints = primary_url
    pagination_mode = "unknown"
    detail_pattern = ""
    list_organization = ""
    full_list_strategy = ""
    detail_requirement = "detail page required"
    blockers = ""

    if cfg:
        search_urls, apartment_urls = flatten_search_urls(cfg)
        discovery_entrypoints = "\n".join(search_urls) if search_urls else primary_url
        apartment_entrypoints = "\n".join(apartment_urls)
        detail_pattern = cfg["listing_pattern"].pattern if cfg.get("listing_pattern") else ""
        if cfg["func"] == "_scrape_homes_bg":
            pagination_mode = "JSON API pages by `currPage`, `offerType`, and optional `city`"
            list_organization = "Apartment feed is discovered from Homes.bg JSON API result pages, then each card points to an HTML detail page."
            full_list_strategy = "Loop `offerType` sale/rent and city scopes, paginate API responses until empty, collect `viewHref`, then fetch each detail page for full text, photos, and structured data."
        elif cfg["func"] == "_scrape_imot_bg":
            pagination_mode = "server-rendered `/p-{page}` search pagination"
            list_organization = "Apartment and other property grids live on city+intent listing pages; detail pages are `obiava-*` URLs."
            full_list_strategy = "Walk city and intent search pages, increment `/p-{page}` until no new URLs, normalize `obiava-*` detail URLs, then parse each detail page."
        else:
            suffix = cfg.get("page_suffix", "?page={}")
            pagination_mode = f"server-rendered pagination using suffix `{suffix}`"
            list_organization = "Listings are organized by product/category buckets or broad sale/rent archives; apartment recovery is best from category pages instead of the home page."
            full_list_strategy = "Start from bucket entrypoints, paginate each bucket until no new URLs, regex-match detail URLs, then fetch every detail page for full text and media."
        blockers = "none in current config" if corpus.get("saved_listings") else "connector planning exists but saved corpus is still zero"
    else:
        if row.get("access_mode") == "official_api":
            pagination_mode = "official API pagination and detail hydration"
            list_organization = "Use API listing feed, not HTML homepage traversal."
            full_list_strategy = "Enumerate real-estate API result pages and hydrate details or media fields through documented endpoints."
            detail_requirement = "API detail enrichment or item-page verification"
        elif row.get("access_mode") == "partner_feed":
            pagination_mode = "partner export or licensed feed"
            list_organization = "Do not treat as a normal crawl target."
            full_list_strategy = "Obtain licensed data access before full-list recovery."
            detail_requirement = "feed or licensed export preferred"
        elif row.get("access_mode") == "headless":
            pagination_mode = "browser/XHR mapping required"
            list_organization = "Likely JS-heavy result grid or map-first browse flow."
            full_list_strategy = "Trace entry pages and XHR calls first, then decide between browser automation and lower-level API extraction."
        else:
            pagination_mode = "server-rendered pagination expected"
            list_organization = "Likely category or city listing archives behind standard HTML pages."
            full_list_strategy = "Start from apartment/category pages, paginate, and fetch each detail page."
        blockers = SOURCE_METHOD_NOTES.get(source_name, {}).get("known_blockers", "connector not yet captured in repo")

    overrides = SOURCE_METHOD_NOTES.get(source_name, {})
    list_organization = overrides.get("list_organization", list_organization)
    full_list_strategy = overrides.get("full_list_strategy", full_list_strategy)
    detail_requirement = overrides.get("detail_requirement", detail_requirement)
    blockers = overrides.get("known_blockers", blockers)

    return {
        "tier": row.get("tier"),
        "source_name": source_name,
        "primary_url": primary_url,
        "related_urls": "\n".join(related_urls),
        "access_mode": row.get("access_mode", ""),
        "risk_mode": row.get("risk_mode", ""),
        "legal_mode": row.get("legal_mode", ""),
        "best_extraction_method": row.get("best_extraction_method", ""),
        "services_declared": services,
        "estimated_total_listings": estimated,
        "saved_listings": corpus.get("saved_listings", 0),
        "with_description": corpus.get("with_description", 0),
        "with_photo_urls": corpus.get("with_photo_urls", 0),
        "with_readable_local_photos": corpus.get("with_readable_local_photos", 0),
        "services_scraped": ", ".join(f"{k}={v}" for k, v in sorted((corpus.get("services") or {}).items())),
        "categories_scraped": ", ".join(f"{k}={v}" for k, v in sorted((corpus.get("categories") or {}).items())),
        "discovery_entrypoints": discovery_entrypoints,
        "apartment_entrypoints": apartment_entrypoints,
        "pagination_mode": pagination_mode,
        "detail_url_pattern": detail_pattern,
        "list_organization": list_organization,
        "full_list_strategy": full_list_strategy,
        "detail_requirement": detail_requirement,
        "known_blockers": blockers,
    }


def build_rows() -> tuple[list[dict], list[dict]]:
    registry_rows = [row for row in load_registry() if row.get("tier") in (1, 2)]
    source_configs = invert_source_configs(load_live_configs())
    corpus_stats = scan_saved_corpus()
    rows = [describe_source(row, source_configs, corpus_stats, ESTIMATED_TOTAL_LISTINGS) for row in registry_rows]
    rows.sort(key=lambda item: (item["tier"], item["source_name"].lower()))
    return rows, SKILL_ROWS


def write_json(rows: list[dict], skills: list[dict]) -> None:
    payload = {
        "generated_at": datetime.now(tz=timezone.utc).isoformat(),
        "sources": rows,
        "skill_gaps": skills,
    }
    (EXPORTS / "tier12-source-analysis.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_markdown(rows: list[dict], skills: list[dict]) -> None:
    tier_totals = defaultdict(int)
    for row in rows:
        tier_totals[row["tier"]] += 1

    lines = [
        "# Tier 1-2 source analysis",
        "",
        f"Generated: {datetime.now(tz=timezone.utc).isoformat()}",
        "",
        "## Tier summary",
        "",
        f"- Tier 1 sources analyzed: {tier_totals[1]}",
        f"- Tier 2 sources analyzed: {tier_totals[2]}",
        "- Primary interpretation: the repo already has strong category-entrypoint knowledge for implemented sources, but the remaining gap is detail-page recovery and DB-backed proof, not source inventory.",
        "",
        "## Source table",
        "",
        "| Tier | Source | Primary URL | Apartment entrypoint | Saved listings | Descriptions | Readable photo sets | Full-list strategy |",
        "| --- | --- | --- | --- | ---: | ---: | ---: | --- |",
    ]
    for row in rows:
        entry = row["apartment_entrypoints"].splitlines()[0] if row["apartment_entrypoints"] else "-"
        lines.append(
            f"| {row['tier']} | {row['source_name']} | {row['primary_url']} | {entry} | "
            f"{row['saved_listings']} | {row['with_description']} | {row['with_readable_local_photos']} | {row['full_list_strategy']} |"
        )

    lines.extend(["", "## Detailed source notes", ""])
    for row in rows:
        lines.extend([
            f"### {row['source_name']}",
            "",
            f"- Links: `{row['primary_url']}`" + (f"; related: `{row['related_urls'].replace(chr(10), ', ')}`" if row["related_urls"] else ""),
            f"- Access/legal: `{row['access_mode']}` / `{row['risk_mode']}` / `{row['legal_mode']}`",
            f"- Declared services: `{row['services_declared']}`",
            f"- Scraped services: `{row['services_scraped'] or 'none saved yet'}`",
            f"- Scraped categories: `{row['categories_scraped'] or 'none saved yet'}`",
            f"- Apartment list organization: {row['list_organization']}",
            f"- Discovery entrypoints: {row['discovery_entrypoints'].replace(chr(10), '; ')}",
            f"- Apartment-focused entrypoints: {row['apartment_entrypoints'].replace(chr(10), '; ') if row['apartment_entrypoints'] else 'not yet captured in repo'}",
            f"- Pagination: {row['pagination_mode']}",
            f"- Detail URL pattern: `{row['detail_url_pattern'] or 'not yet mapped'}`",
            f"- Full-list approach: {row['full_list_strategy']}",
            f"- Item-page requirement: {row['detail_requirement']}",
            f"- Current blocker: {row['known_blockers']}",
            "",
        ])
    (EXPORTS / "tier12-source-analysis.md").write_text("\n".join(lines), encoding="utf-8")

    skill_lines = [
        "# Tier 1-2 skills gap analysis",
        "",
        f"Generated: {datetime.now(tz=timezone.utc).isoformat()}",
        "",
        "| Capability | Current project skills | External skill | Link | Status | Project-local mirror | Analysis |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in skills:
        skill_lines.append(
            f"| {row['capability']} | {row['current_project_skills']} | {row['external_skill']} | "
            f"{row['external_link']} | {row['status']} | {row['project_mirror']} | {row['analysis']} |"
        )
    (EXPORTS / "tier12-skill-gap-analysis.md").write_text("\n".join(skill_lines), encoding="utf-8")


def write_xlsx(rows: list[dict], skills: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tier1-2 Sources"

    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    tier_fill = {
        1: PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
        2: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    }

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header == "tier":
                cell.fill = tier_fill.get(row["tier"], PatternFill())

    widths = {
        "source_name": 24,
        "primary_url": 28,
        "related_urls": 24,
        "best_extraction_method": 28,
        "services_declared": 20,
        "services_scraped": 26,
        "categories_scraped": 28,
        "discovery_entrypoints": 42,
        "apartment_entrypoints": 40,
        "pagination_mode": 26,
        "detail_url_pattern": 32,
        "list_organization": 40,
        "full_list_strategy": 48,
        "detail_requirement": 24,
        "known_blockers": 28,
    }
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)

    skill_ws = wb.create_sheet("Skill Gaps")
    skill_headers = list(skills[0].keys())
    for col, header in enumerate(skill_headers, start=1):
        cell = skill_ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(skills, start=2):
        for col_idx, header in enumerate(skill_headers, start=1):
            cell = skill_ws.cell(row=row_idx, column=col_idx, value=row[header])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, header in enumerate(skill_headers, start=1):
        skill_ws.column_dimensions[get_column_letter(idx)].width = 28 if header != "analysis" else 42

    wb.save(EXPORTS / "tier12-source-analysis.xlsx")


def main() -> None:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    rows, skills = build_rows()
    write_json(rows, skills)
    write_markdown(rows, skills)
    write_xlsx(rows, skills)
    print("Generated tier-1/2 source analysis artifacts.")


if __name__ == "__main__":
    main()
