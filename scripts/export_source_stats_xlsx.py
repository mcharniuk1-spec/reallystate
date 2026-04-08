from __future__ import annotations

import os
from pathlib import Path

from bgrealestate.connectors.factory import marketplace_sources
from bgrealestate.connectors.legal import default_primary_endpoint, derive_default_legal_rule
from bgrealestate.source_registry import SourceRegistry
from bgrealestate.stats.source_stats import SourceStatRow, fetch_source_stats


def main() -> int:
    url = os.getenv("DATABASE_URL")

    out_path = Path(os.getenv("OUT_XLSX", "docs/exports/source-stats.xlsx"))
    out_path.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]

    rows: list[SourceStatRow]
    mode: str
    if url:
        from sqlalchemy import create_engine

        engine = create_engine(url, pool_pre_ping=True)
        rows = fetch_source_stats(engine)
        mode = "live-db"
    else:
        # Fallback for environments where DB is intentionally unavailable.
        # Keeps reporting artifacts reproducible while clearly signaling mode.
        registry = SourceRegistry.from_file(Path("data/source_registry.json"))
        rows = []
        for entry in marketplace_sources(registry):
            rule = derive_default_legal_rule(entry)
            ep = default_primary_endpoint(entry)
            rows.append(
                SourceStatRow(
                    source_name=entry.source_name,
                    tier=entry.tier,
                    legal_mode=entry.legal_mode,
                    canonical_listings=0,
                    raw_captures=0,
                    with_description=0,
                    with_photos=0,
                    photo_coverage_pct=0.0,
                    intent_sale=0,
                    intent_rent=0,
                    intent_str=0,
                    intent_auction=0,
                    category_apartment=0,
                    category_house=0,
                    category_land=0,
                    category_commercial=0,
                    has_legal_rule=bool(rule),
                    has_endpoint=ep is not None,
                )
            )
        mode = "bootstrap-no-db"

    wb = Workbook()
    ws = wb.active
    ws.title = "source_stats"

    header = [
        "source_name",
        "tier",
        "legal_mode",
        "canonical_listings",
        "raw_captures",
        "with_description",
        "with_photos",
        "photo_coverage_pct",
        "intent_sale",
        "intent_rent",
        "intent_str",
        "intent_auction",
        "category_apartment",
        "category_house",
        "category_land",
        "category_commercial",
        "has_legal_rule",
        "has_endpoint",
    ]
    ws.append(header)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in rows:
        ws.append([
            r.source_name,
            r.tier,
            r.legal_mode,
            r.canonical_listings,
            r.raw_captures,
            r.with_description,
            r.with_photos,
            r.photo_coverage_pct,
            r.intent_sale,
            r.intent_rent,
            r.intent_str,
            r.intent_auction,
            r.category_apartment,
            r.category_house,
            r.category_land,
            r.category_commercial,
            "Y" if r.has_legal_rule else "N",
            "Y" if r.has_endpoint else "N",
        ])

    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    wb.save(out_path)
    print(f"wrote {out_path} ({len(rows)} rows, mode={mode})")
    if mode != "live-db":
        print("note: generated without DATABASE_URL; counts are bootstrap defaults (0) until DB sync runs.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
