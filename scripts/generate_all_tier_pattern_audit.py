#!/usr/bin/env python3
"""Generate all-tier pattern audit artifacts for current source coverage."""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from openpyxl import Workbook  # type: ignore[import-untyped]  # noqa: E402
from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]  # noqa: E402
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]  # noqa: E402

from bgrealestate.connectors.factory import build_connector  # noqa: E402
from bgrealestate.source_registry import SourceRegistry  # noqa: E402


REGISTRY_PATH = ROOT / "data" / "source_registry.json"
PATTERN_STATUS_PATH = ROOT / "docs" / "exports" / "tier12-pattern-status.json"
SECTIONS_PATH = ROOT / "data" / "scrape_patterns" / "regions" / "varna" / "sections.json"
OUTPUT_JSON = ROOT / "docs" / "exports" / "all-tier-source-pattern-audit-2026-04-30.json"
OUTPUT_MD = ROOT / "docs" / "exports" / "all-tier-source-pattern-audit-2026-04-30.md"
OUTPUT_XLSX = ROOT / "docs" / "exports" / "all-tier-source-pattern-audit-2026-04-30.xlsx"
PATTERN_CANDIDATES_JSON = ROOT / "data" / "scrape_patterns" / "pattern_candidates" / "all-tier-unpatterned-source-patterns.json"
PATTERN_CANDIDATES_MD = ROOT / "docs" / "exports" / "all-tier-unpatterned-source-patterns-2026-04-30.md"
FIXTURES_ROOT = ROOT / "tests" / "fixtures"
SCRAPED_ROOT = ROOT / "data" / "scraped"

ACTION1_SOURCES = {
    "Address.bg",
    "BulgarianProperties",
    "Homes.bg",
    "imot.bg",
    "LUXIMMO",
    "property.bg",
    "SUPRIMMO",
}
PATTERNED_SECONDARY_REVIEW = {"OLX.bg", "Bazar.bg", "Yavlena"}

SOURCE_SLUGS = {
    "Address.bg": "address_bg",
    "alo.bg": "alo_bg",
    "BulgarianProperties": "bulgarianproperties",
    "Homes.bg": "homes_bg",
    "imot.bg": "imot_bg",
    "imoti.net": "imoti_net",
    "LUXIMMO": "luximmo",
    "OLX.bg": "olx_bg",
    "property.bg": "property_bg",
    "SUPRIMMO": "suprimmo",
    "ApartmentsBulgaria.com": "apartmentsbulgaria",
    "Bazar.bg": "bazar_bg",
    "Domaza": "domaza",
    "Holding Group Real Estate": "holding_group",
    "Home2U": "home2u",
    "Imoteka.bg": "imoteka",
    "Imoti.info": "imoti_info",
    "Indomio.bg": "indomio",
    "Lions Group": "lions_group",
    "Pochivka.bg": "pochivka",
    "realestates.bg": "realestates",
    "Realistimo": "realistimo",
    "Rentica.bg": "rentica",
    "Svobodni-kvartiri.com": "svobodni_kvartiri",
    "Unique Estates": "unique_estates",
    "Vila.bg": "vila",
    "Yavlena": "yavlena",
}

PATTERN_KIND_BY_ACCESS = {
    "html": "html_list_detail_gallery",
    "headless": "headless_list_detail_gallery",
    "partner_feed": "partner_feed_or_vendor_contract",
    "licensed_data": "licensed_data_ingest",
    "official_api": "official_api_overlay",
    "manual_consent_only": "manual_or_consent_flow",
}

STATUS_EXPLANATIONS = {
    "Patterned": "Strict pattern proof exists from a saved sample item with local gallery files and core fields.",
    "without_sample_product_capture": "Routes or parser evidence may exist, but the repo does not yet have one saved full product item proving detail + gallery capture.",
    "without_authorized_pattern": "The source is blocked by legal/licensing review before live pattern promotion.",
    None: "No tier12 pattern row exists; use legal/access pattern only.",
}

MANUAL_CONTENT_NOTES = {
    "Holding Group Real Estate": "Agency website with sale/rent search widgets, city filters, and detail pages carrying area, beds/baths, and city labels in public content.",
    "Home2U": "Agency portal with `/properties/` search, explicit sale/rent filters, city landing pages, and descriptive listing cards in public HTML.",
    "Rentica.bg": "Rent-first Varna agency catalog with numeric offer detail pages, district pages, and explicit long-term rental descriptions.",
    "Realistimo": "Large SSR portal with city-level buy/rent landing pages, broad property taxonomy, and explicit city result counts in public pages.",
    "Unique Estates": "Luxury agency portal with explicit buy/rent navigation, editorial landing pages, and premium listing detail pages.",
    "Vila.bg": "Hospitality catalog oriented around villa/guest-house short-term stays with region/occasion pages rather than classic property-sale inventory.",
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def source_slug(source_name: str) -> str | None:
    return SOURCE_SLUGS.get(source_name)


def source_dir(source_name: str) -> Path | None:
    slug = source_slug(source_name)
    if not slug:
        return None
    return SCRAPED_ROOT / slug


def fixture_cases(source_name: str) -> list[str]:
    slug = source_slug(source_name)
    if not slug:
        return []
    fixture_dir = FIXTURES_ROOT / slug
    if not fixture_dir.exists():
        return []
    return sorted(path.name for path in fixture_dir.iterdir() if path.is_dir())


def section_rows_by_source() -> dict[str, list[dict[str, Any]]]:
    if not SECTIONS_PATH.exists():
        return {}
    rows = load_json(SECTIONS_PATH).get("sections", [])
    out: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        out[row["source_name"]].append(row)
    return dict(out)


def listing_counts_for(source_name: str) -> dict[str, Any]:
    out = {
        "saved_items": 0,
        "services": defaultdict(int),
        "categories": defaultdict(int),
        "buckets": defaultdict(int),
        "full_gallery_items": 0,
    }
    sdir = source_dir(source_name)
    if not sdir:
        return out
    listing_dir = sdir / "listings"
    if not listing_dir.exists():
        return out
    for path in listing_dir.glob("*.json"):
        try:
            row = load_json(path)
        except Exception:
            continue
        out["saved_items"] += 1
        out["services"][str(row.get("listing_intent") or "unknown")] += 1
        out["categories"][str(row.get("property_category") or "unknown")] += 1
        bucket = classify_bucket(row)
        out["buckets"][bucket] += 1
        remote = len(row.get("image_urls") or [])
        local = len(row.get("local_image_files") or [])
        if local and local >= remote:
            out["full_gallery_items"] += 1
    out["services"] = dict(out["services"])
    out["categories"] = dict(out["categories"])
    out["buckets"] = dict(out["buckets"])
    return out


def classify_bucket(row: dict[str, Any]) -> str:
    intent = str(row.get("listing_intent") or "")
    category = str(row.get("property_category") or "")
    commercial = category in {"office", "commercial", "shop", "industrial", "warehouse"}
    if intent in {"sale", "new_build", "auction_sale"}:
        return "buy_commercial" if commercial else "buy_personal"
    if intent in {"long_term_rent", "short_term_rent"}:
        return "rent_commercial" if commercial else "rent_personal"
    return "unknown"


def connector_name(source_name: str, registry: SourceRegistry) -> str:
    try:
        return build_connector(source_name, registry).__class__.__name__
    except Exception as exc:
        return f"connector_error:{type(exc).__name__}"


def route_evidence(source_name: str, section_rows: list[dict[str, Any]], registry_entry: dict[str, Any]) -> dict[str, Any]:
    primary = registry_entry.get("primary_url") or ""
    related = list(registry_entry.get("related_urls") or [])
    entry_urls = sorted({url for row in section_rows for url in row.get("entry_urls") or []})
    if entry_urls:
        return {
            "route_level": "entry_urls_saved",
            "entry_urls": entry_urls,
            "segment_count": len(section_rows),
        }
    if related:
        return {
            "route_level": "related_urls_only",
            "entry_urls": related[:8],
            "segment_count": 0,
        }
    if primary:
        return {
            "route_level": "primary_url_only",
            "entry_urls": [primary],
            "segment_count": 0,
        }
    return {"route_level": "no_route_evidence", "entry_urls": [], "segment_count": 0}


def action_lane(source_name: str, tier: int) -> str:
    if source_name in ACTION1_SOURCES:
        return "action1_owned"
    if source_name in PATTERNED_SECONDARY_REVIEW:
        return "patterned_secondary_review"
    if tier in {1, 2}:
        return "unpatterned_focus"
    return "legal_or_partner_pattern_only"


def proposed_pattern_kind(entry: dict[str, Any]) -> str:
    return PATTERN_KIND_BY_ACCESS.get(str(entry.get("access_mode") or ""), "manual_analysis_required")


def issue_rows_for(
    entry: dict[str, Any],
    pattern_row: dict[str, Any] | None,
    route_row: dict[str, Any],
    fixtures: list[str],
    connector: str,
    route_mismatch_count: int,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_name = entry["source_name"]
    legal_mode = str(entry.get("legal_mode") or "")
    access_mode = str(entry.get("access_mode") or "")
    pattern_status = pattern_row.get("pattern_status") if pattern_row else None

    if legal_mode in {"legal_review_required", "licensing_required", "official_partner_or_vendor_only", "consent_or_manual_only"}:
        rows.append(
            issue_row(
                source_name,
                "LEGAL-01",
                "Legal or contract gate blocks promotion",
                f"Current legal_mode={legal_mode}, access_mode={access_mode}.",
                [
                    "Validate whether a public scraping pattern is allowed at all.",
                    "If not, switch the source to partner-feed, official API, licensed-data, or manual-only execution.",
                    "Do not attempt live pattern promotion until the legal gate changes.",
                ],
            )
        )

    if pattern_status == "without_sample_product_capture":
        rows.append(
            issue_row(
                source_name,
                "EVID-01",
                "No saved full product sample",
                "The repo does not contain one saved detail-page item with full local gallery proof for this source.",
                [
                    "Capture one legal detail page from a supported bucket.",
                    "Persist raw HTML, normalized listing JSON, and local gallery files.",
                    "Re-run strict pattern audit after the sample exists.",
                ],
            )
        )

    if route_row["route_level"] in {"primary_url_only", "no_route_evidence"}:
        rows.append(
            issue_row(
                source_name,
                "ROUTE-01",
                "Section or list-route discovery is incomplete",
                f"Current route evidence level is {route_row['route_level']}.",
                [
                    "Map buy/rent and residential/commercial landing pages explicitly.",
                    "Persist section entry URLs and pagination rules.",
                    "Add at least one route per supported bucket before scaling.",
                ],
            )
        )

    if route_mismatch_count:
        rows.append(
            issue_row(
                source_name,
                "MAP-01",
                "Saved section routes do not match their segment labels",
                f"{route_mismatch_count} saved section route(s) look inconsistent with buy/rent semantics.",
                [
                    "Review each saved section URL against the intended segment key.",
                    "Correct the section registry before treating the routes as scrape-ready.",
                    "Re-run the audit so pattern candidates point to the right listing surfaces.",
                ],
            )
        )

    if not fixtures and access_mode in {"html", "headless"}:
        rows.append(
            issue_row(
                source_name,
                "FIX-01",
                "No fixture-backed parser evidence",
                "This source has no dedicated fixture cases in tests/fixtures.",
                [
                    "Save at least one representative detail page fixture.",
                    "Add one non-happy-path fixture when the site exposes multiple templates.",
                    "Bind the fixture to a parser test before live promotion.",
                ],
            )
        )

    code_paths = pattern_row.get("code_paths") if pattern_row else []
    has_source_parser = any("::_parse_" in str(path) or "::parse_" in str(path) for path in code_paths or [])
    if connector == "HtmlPortalConnector" and access_mode in {"html", "headless"} and not has_source_parser:
        rows.append(
            issue_row(
                source_name,
                "PARSE-01",
                "Only generic HTML parser is wired",
                "The source currently relies on the generic JSON-LD/og:image parser path.",
                [
                    "Validate whether generic extraction is enough for price, area, rooms, address, and full gallery.",
                    "If not, add a source-specific parser or runtime adapter.",
                    "Prove the parser on at least two materially different property pages.",
                ],
            )
        )

    if access_mode == "headless":
        rows.append(
            issue_row(
                source_name,
                "RUNTIME-01",
                "Headless/browser runtime likely required",
                "Registry marks this source as headless-driven, so HTML-only assumptions are weak.",
                [
                    "Confirm whether SSR HTML is sufficient or whether browser state is required.",
                    "Persist a browser-safe route and extraction approach.",
                    "Add a non-interactive fallback only if it is stable.",
                ],
            )
        )

    return rows


def issue_row(source_name: str, issue_id: str, title: str, detail: str, steps: list[str]) -> dict[str, str]:
    return {
        "source_name": source_name,
        "issue_id": issue_id,
        "title": title,
        "detail": detail,
        "steps": " ; ".join(f"{i + 1}. {step}" for i, step in enumerate(steps)),
    }


def issue_steps_markdown(steps: str) -> str:
    return "<br>".join(part.strip() for part in steps.split(" ; ") if part.strip())


def content_summary(entry: dict[str, Any]) -> str:
    listing_types = ", ".join(entry.get("listing_types") or [])
    family = entry.get("source_family") or "unknown"
    manual = MANUAL_CONTENT_NOTES.get(entry["source_name"])
    if manual:
        return f"{family}; listing_types={listing_types}. {manual}"
    return f"{family}; listing_types={listing_types}."


def sample_evidence_summary(pattern_row: dict[str, Any] | None) -> str:
    if not pattern_row:
        return "No tier12 sample row."
    sample = pattern_row.get("sample")
    if not sample:
        return "No saved sample item."
    return (
        f"{sample.get('reference_id')} | photos {sample.get('local_photo_count')}/{sample.get('remote_photo_count')} "
        f"| desc chars {sample.get('description_chars')} | structured fields {sample.get('structured_fields_count')}"
    )


def universality_rows(pattern_rows: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    fixture_map = {
        "OLX.bg": fixture_cases("OLX.bg"),
        "Bazar.bg": fixture_cases("Bazar.bg"),
        "Yavlena": fixture_cases("Yavlena"),
    }
    for source_name in ["OLX.bg", "Bazar.bg", "Yavlena"]:
        pattern_row = pattern_rows[source_name]
        listing_stats = listing_counts_for(source_name)
        services = sorted(listing_stats["services"].keys())
        categories = sorted(listing_stats["categories"].keys())
        fixtures = fixture_map[source_name]
        if source_name == "OLX.bg":
            status = "broad_schema_but_not_fully_proven_universal"
            reason = (
                "API-backed parser covers sale/rent plus apartment/house/land/office in saved corpus, "
                "but fixture proof is limited to basic + missing-price detail shapes and no new_build-specific evidence exists."
            )
        elif source_name == "Bazar.bg":
            status = "not_proven_universal"
            reason = (
                "Saved corpus spans sale/rent and some office rows, but fixtures only prove apartment and land; "
                "rent/commercial layouts are not directly fixture-backed."
            )
        else:
            status = "not_proven_universal"
            reason = (
                "Saved corpus spans sale houses/land/offices and some rent, but fixtures only prove one sale and one rent case; "
                "gallery depth is also only one image in the best sample."
            )
        rows.append(
            {
                "source_name": source_name,
                "parser_method": pattern_row.get("method"),
                "status": status,
                "reason": reason,
                "saved_items": listing_stats["saved_items"],
                "services": services,
                "categories": categories,
                "fixture_cases": fixtures,
                "sample_evidence": sample_evidence_summary(pattern_row),
            }
        )
    return rows


def supported_segments(entry: dict[str, Any]) -> list[str]:
    listing_types = set(entry.get("listing_types") or [])
    segments: list[str] = []
    if listing_types & {"sale", "new_build", "auction_sale", "land"}:
        segments.extend(["buy_personal", "buy_commercial"])
    if listing_types & {"long_term_rent", "short_term_rent"}:
        segments.extend(["rent_personal", "rent_commercial"])
    return segments or ["unknown"]


def route_segment_mismatch(segment_key: str, entry_url: str | None) -> bool:
    if not entry_url:
        return False
    url = entry_url.lower()
    rent_tokens = ["naem", "naemi", "rent", "pod-naem", "for-rent"]
    buy_tokens = ["prodaj", "sale", "buy", "for-sale"]
    if segment_key.startswith("buy_") and any(token in url for token in rent_tokens):
        return True
    if segment_key.startswith("rent_") and any(token in url for token in buy_tokens):
        return True
    return False


def section_pattern_hypotheses(
    entry: dict[str, Any],
    route_row: dict[str, Any],
    source_sections: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if source_sections:
        rows: list[dict[str, Any]] = []
        for section in source_sections:
            entry_urls = list(section.get("entry_urls") or [])
            entry_url = entry_urls[0] if entry_urls else None
            mismatch = route_segment_mismatch(str(section.get("segment_key") or "unknown"), entry_url)
            rows.append(
                {
                    "segment_key": section.get("segment_key") or "unknown",
                    "vertical_key": section.get("vertical_key") or "all",
                    "entry_url": entry_url,
                    "route_status": "entry_urls_saved_mismatch" if mismatch else "entry_urls_saved" if entry_urls else "needs_route_discovery",
                    "notes": "Saved route appears mismatched with the segment label and needs correction." if mismatch else "Section route is persisted in the Varna section registry; still needs strict item-level proof." if entry_urls else "Section exists logically but lacks a saved route.",
                }
            )
        return rows

    return [
        {
            "segment_key": segment,
            "vertical_key": "all",
            "entry_url": None if route_row["route_level"] == "no_route_evidence" else (route_row.get("entry_urls") or [None])[0],
            "route_status": "needs_route_discovery" if route_row["route_level"] == "no_route_evidence" else route_row["route_level"],
            "notes": "No saved section route exists yet; derive city/operation landing pages before parser promotion.",
        }
        for segment in supported_segments(entry)
    ]


def candidate_pattern(entry: dict[str, Any], row: dict[str, Any], source_sections: list[dict[str, Any]]) -> dict[str, Any]:
    legal_mode = str(entry.get("legal_mode") or "")
    public_scrape_allowed = legal_mode == "public_crawl_with_review"
    section_patterns = section_pattern_hypotheses(entry, row["route_evidence"], source_sections)
    return {
        "source_name": row["source_name"],
        "tier": row["tier"],
        "status": row["current_pattern_status"] or "n/a",
        "promotion_gate": "strict_sample_and_gallery_proof" if public_scrape_allowed else "legal_or_partner_gate",
        "pattern_scope": "public_scrape_candidate" if public_scrape_allowed else "restricted_access_pattern",
        "proposed_pattern_kind": row["proposed_pattern_kind"],
        "connector_class": row["connector_class"],
        "access_mode": row["access_mode"],
        "legal_mode": row["legal_mode"],
        "source_level_pattern": {
            "primary_url": row["primary_url"],
            "related_urls": row["related_urls"],
            "route_level": row["route_evidence"]["route_level"],
            "content_summary": row["content_summary"],
            "supported_listing_types": row["listing_types"],
        },
        "section_level_patterns": section_patterns,
        "list_page_pattern": {
            "mode": "html_listing_grid" if row["access_mode"] == "html" else "browser_or_ssr_listing_grid" if row["access_mode"] == "headless" else "non_public_feed_or_api",
            "proof_status": "saved_route_exists" if row["route_evidence"]["entry_urls"] else "missing_saved_route",
            "known_entry_urls": row["route_evidence"]["entry_urls"],
        },
        "detail_page_pattern": {
            "mode": "detail_html_to_listing_json",
            "proof_status": "saved_sample_exists" if row["sample_evidence"] != "No saved sample item." else "missing_saved_sample",
            "required_fields": [
                "title",
                "price or price_status",
                "city or address",
                "description",
                "image_urls",
                "local_image_files",
                "at least two structured fields",
            ],
        },
        "gallery_pattern": {
            "mode": "full_reachable_gallery_to_local_files",
            "proof_status": "full_gallery_sample_exists" if row["full_gallery_items"] else "no_full_gallery_item_saved",
            "current_full_gallery_items": row["full_gallery_items"],
        },
        "fixture_cases": row["fixture_cases"],
        "code_paths": row["code_paths"],
        "issues": row["issues"],
    }


def build_registry() -> dict[str, Any]:
    registry = SourceRegistry.from_file(REGISTRY_PATH)
    source_rows = load_json(REGISTRY_PATH)["sources"]
    pattern_rows = {row["source_name"]: row for row in load_json(PATTERN_STATUS_PATH).get("sources", [])}
    section_rows = section_rows_by_source()

    output_sources: list[dict[str, Any]] = []
    issue_rows: list[dict[str, str]] = []

    candidate_patterns: list[dict[str, Any]] = []

    for entry in sorted(source_rows, key=lambda row: (row.get("tier", 99), row["source_name"].lower())):
        source_name = entry["source_name"]
        pattern_row = pattern_rows.get(source_name)
        fixtures = fixture_cases(source_name)
        connector = connector_name(source_name, registry)
        routes = route_evidence(source_name, section_rows.get(source_name, []), entry)
        listing_stats = listing_counts_for(source_name)
        pattern_status = pattern_row.get("pattern_status") if pattern_row else None
        route_mismatch_count = sum(
            1
            for section in section_rows.get(source_name, [])
            if route_segment_mismatch(str(section.get("segment_key") or "unknown"), (section.get("entry_urls") or [None])[0])
        )
        issues = issue_rows_for(entry, pattern_row, routes, fixtures, connector, route_mismatch_count)
        issue_rows.extend(issues)

        output_row = {
                "source_name": source_name,
                "tier": entry.get("tier"),
                "action_lane": action_lane(source_name, int(entry.get("tier") or 0)),
                "source_family": entry.get("source_family"),
                "primary_url": entry.get("primary_url"),
                "related_urls": entry.get("related_urls") or [],
                "listing_types": entry.get("listing_types") or [],
                "legal_mode": entry.get("legal_mode"),
                "risk_mode": entry.get("risk_mode"),
                "access_mode": entry.get("access_mode"),
                "current_pattern_status": pattern_status,
                "current_pattern_issue": pattern_row.get("pattern_issue") if pattern_row else "No tier12 pattern row.",
                "status_explanation": STATUS_EXPLANATIONS.get(pattern_status),
                "proposed_pattern_kind": proposed_pattern_kind(entry),
                "content_summary": content_summary(entry),
                "connector_class": connector,
                "fixture_cases": fixtures,
                "fixture_count": len(fixtures),
                "route_evidence": routes,
                "route_mismatch_count": route_mismatch_count,
                "sample_evidence": sample_evidence_summary(pattern_row),
                "saved_items": listing_stats["saved_items"],
                "saved_services": listing_stats["services"],
                "saved_categories": listing_stats["categories"],
                "saved_buckets": listing_stats["buckets"],
                "full_gallery_items": listing_stats["full_gallery_items"],
                "website_total_active": (pattern_row or {}).get("website_total_active"),
                "count_status": (pattern_row or {}).get("count_status"),
                "recent_status": (pattern_row or {}).get("recent_status"),
                "varna_status": (pattern_row or {}).get("varna_status"),
                "pattern_method": (pattern_row or {}).get("method"),
                "code_paths": (pattern_row or {}).get("code_paths") or [],
                "issues": issues,
            }
        output_sources.append(output_row)
        if output_row["action_lane"] == "unpatterned_focus":
            candidate_patterns.append(candidate_pattern(entry, output_row, section_rows.get(source_name, [])))

    return {
        "generated_at": "2026-04-30",
        "action1_sources": sorted(ACTION1_SOURCES),
        "patterned_secondary_review": sorted(PATTERNED_SECONDARY_REVIEW),
        "sources": output_sources,
        "issues": issue_rows,
        "universality": universality_rows(pattern_rows),
        "candidate_patterns": candidate_patterns,
    }


def write_json(payload: dict[str, Any]) -> None:
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def format_section_routes(routes: dict[str, Any]) -> str:
    urls = routes.get("entry_urls") or []
    if not urls:
        return "none"
    return "; ".join(urls[:6])


def write_markdown(payload: dict[str, Any]) -> None:
    sources = payload["sources"]
    universality = payload["universality"]
    candidate_patterns = payload["candidate_patterns"]
    unpatterned = [row for row in sources if row["current_pattern_status"] != "Patterned" and row["action_lane"] == "unpatterned_focus"]
    legal_only = [row for row in sources if row["action_lane"] == "legal_or_partner_pattern_only"]

    lines: list[str] = []
    lines.append("# All-tier source pattern audit (2026-04-30)")
    lines.append("")
    lines.append("## Scope")
    lines.append("")
    lines.append("- FACT: OpenClaw Action1 owns the seven priority patterned sources and is excluded from this run's repair scope.")
    lines.append("- FACT: This audit focuses on unpatterned sources across all tiers plus universality checks for patterned non-Action1 sources.")
    lines.append("- INTERPRETATION: not every source should become a public scraper; some should stay partner-feed, official-API, or manual-only by design.")
    lines.append("")
    lines.append("## Summary table")
    lines.append("")
    lines.append("| Source | Tier | Lane | Current status | Proposed pattern kind | Connector | Fixtures | Route evidence | Main blocker |")
    lines.append("|---|---:|---|---|---|---|---:|---|---|")
    for row in sources:
        main_blocker = row["issues"][0]["issue_id"] if row["issues"] else "none"
        lines.append(
            f"| {row['source_name']} | {row['tier']} | {row['action_lane']} | "
            f"`{row['current_pattern_status'] or 'n/a'}` | `{row['proposed_pattern_kind']}` | "
            f"`{row['connector_class']}` | {row['fixture_count']} | `{row['route_evidence']['route_level']}` | `{main_blocker}` |"
        )

    lines.append("")
    lines.append("## Unpatterned tier-1/2 sources")
    lines.append("")
    for row in unpatterned:
        lines.append(f"### {row['source_name']}")
        lines.append("")
        lines.append("- FACT:")
        lines.append(f"  - tier={row['tier']}, family={row['source_family']}, legal_mode={row['legal_mode']}, access_mode={row['access_mode']}")
        lines.append(f"  - listing_types={', '.join(row['listing_types']) or 'none'}")
        lines.append(f"  - connector={row['connector_class']}, fixture_cases={row['fixture_cases'] or 'none'}")
        lines.append(f"  - route_evidence={row['route_evidence']['route_level']}, entry_urls={format_section_routes(row['route_evidence'])}")
        lines.append(f"  - current_pattern_status=`{row['current_pattern_status']}`; issue={row['current_pattern_issue']}")
        lines.append("- INTERPRETATION:")
        lines.append(f"  - proposed_pattern_kind=`{row['proposed_pattern_kind']}`")
        lines.append(f"  - content summary: {row['content_summary']}")
        lines.append("- GAP:")
        lines.append(f"  - sample evidence: {row['sample_evidence']}")
        if row["website_total_active"]:
            lines.append(
                f"  - website_total_active={row['website_total_active'].get('value')} "
                f"({row['website_total_active'].get('kind')})"
            )
        lines.append("")
        if row["issues"]:
            lines.append("| Issue ID | Problem | Detail | Steps |")
            lines.append("|---|---|---|---|")
            for issue in row["issues"]:
                lines.append(
                    f"| `{issue['issue_id']}` | {issue['title']} | {issue['detail']} | {issue_steps_markdown(issue['steps'])} |"
                )
            lines.append("")

    lines.append("## Durable candidate patterns for unpatterned tier-1/2 sources")
    lines.append("")
    for row in candidate_patterns:
        lines.append(f"### {row['source_name']}")
        lines.append("")
        lines.append("- FACT:")
        lines.append(f"  - pattern_scope={row['pattern_scope']}, promotion_gate={row['promotion_gate']}")
        lines.append(f"  - connector={row['connector_class']}, access_mode={row['access_mode']}, legal_mode={row['legal_mode']}")
        lines.append("- INTERPRETATION:")
        lines.append(f"  - source-level pattern: `{row['proposed_pattern_kind']}`")
        lines.append(f"  - list-page mode: `{row['list_page_pattern']['mode']}`")
        lines.append(f"  - detail-page proof status: `{row['detail_page_pattern']['proof_status']}`")
        lines.append(f"  - gallery proof status: `{row['gallery_pattern']['proof_status']}`")
        lines.append("- Known section hypotheses:")
        for section in row["section_level_patterns"]:
            lines.append(
                f"  - `{section['segment_key']}` -> {section['entry_url'] or 'route not saved yet'} "
                f"({section['route_status']}; {section['notes']})"
            )
        lines.append("")

    lines.append("## Tier-3 and tier-4 pattern model")
    lines.append("")
    for row in legal_only:
        lines.append(f"### {row['source_name']}")
        lines.append("")
        lines.append("- FACT:")
        lines.append(f"  - tier={row['tier']}, family={row['source_family']}, legal_mode={row['legal_mode']}, access_mode={row['access_mode']}")
        lines.append(f"  - listing_types={', '.join(row['listing_types']) or 'none'}")
        lines.append("- INTERPRETATION:")
        lines.append(f"  - proposed_pattern_kind=`{row['proposed_pattern_kind']}`")
        lines.append(f"  - content summary: {row['content_summary']}")
        lines.append("- GAP:")
        if row["issues"]:
            lines.append(f"  - main blocker: {row['issues'][0]['title']}")
        else:
            lines.append("  - no tier12-style pattern issue row exists; this source should follow its legal/access pattern.")
        lines.append("")

    lines.append("## Patterned non-Action1 universality review")
    lines.append("")
    lines.append("| Source | Status | Saved items | Services seen | Categories seen | Fixture cases | Reason |")
    lines.append("|---|---|---:|---|---|---|---|")
    for row in universality:
        lines.append(
            f"| {row['source_name']} | `{row['status']}` | {row['saved_items']} | "
            f"{', '.join(row['services'])} | {', '.join(row['categories'])} | "
            f"{', '.join(row['fixture_cases'])} | {row['reason']} |"
        )

    lines.append("")
    lines.append("## Conclusions")
    lines.append("")
    lines.append("- FACT: the strongest unpatterned public-scrape candidates already have some route or fixture evidence: `alo.bg`, `Domaza`, and `Home2U`.")
    lines.append("- FACT: several tier-2 sources remain route-poor and fixture-poor; they need route discovery before parser promotion.")
    lines.append("- FACT: tier-3 and much of tier-4 should not be judged by the same public-scrape `Patterned` bar because their legal access model is different.")
    lines.append("- INTERPRETATION: the repo needs a second status axis: `pattern model exists` versus `strict live sample proof exists`.")
    lines.append("- GAP: `OLX.bg`, `Bazar.bg`, and `Yavlena` remain operational but not universally proven across every property/service template on their websites.")
    lines.append("")
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def write_candidate_patterns(payload: dict[str, Any]) -> None:
    PATTERN_CANDIDATES_JSON.parent.mkdir(parents=True, exist_ok=True)
    candidate_payload = {
        "generated_at": payload["generated_at"],
        "sources": payload["candidate_patterns"],
    }
    PATTERN_CANDIDATES_JSON.write_text(json.dumps(candidate_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# All-tier unpatterned source candidate patterns (2026-04-30)",
        "",
        "- FACT: this file persists reusable pattern candidates for current tier-1/2 unpatterned sources.",
        "- INTERPRETATION: it is a durable planning artifact, not a live-proof claim.",
        "",
    ]
    for row in payload["candidate_patterns"]:
        lines.append(f"## {row['source_name']}")
        lines.append("")
        lines.append(f"- Scope: `{row['pattern_scope']}`")
        lines.append(f"- Promotion gate: `{row['promotion_gate']}`")
        lines.append(f"- Proposed pattern: `{row['proposed_pattern_kind']}`")
        lines.append(f"- Connector: `{row['connector_class']}`")
        lines.append(f"- Source route level: `{row['source_level_pattern']['route_level']}`")
        lines.append(f"- Source URLs: {', '.join([u for u in [row['source_level_pattern']['primary_url'], *row['source_level_pattern']['related_urls']] if u]) or 'none'}")
        lines.append("- Section hypotheses:")
        for section in row["section_level_patterns"]:
            lines.append(
                f"  - `{section['segment_key']}` | `{section['route_status']}` | "
                f"{section['entry_url'] or 'route not saved yet'} | {section['notes']}"
            )
        lines.append(f"- Detail requirements: {', '.join(row['detail_page_pattern']['required_fields'])}")
        lines.append(f"- Fixture cases: {', '.join(row['fixture_cases']) or 'none'}")
        lines.append(f"- Code paths: {', '.join(row['code_paths']) or 'none'}")
        lines.append("- Issues:")
        for issue in row["issues"]:
            lines.append(f"  - `{issue['issue_id']}` {issue['title']}: {issue['detail']} :: {issue['steps']}")
        lines.append("")
    PATTERN_CANDIDATES_MD.write_text("\n".join(lines), encoding="utf-8")


def write_xlsx(payload: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    risk_fill = PatternFill("solid", fgColor="FCE4D6")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    hdr = [
        "source_name",
        "tier",
        "action_lane",
        "current_pattern_status",
        "proposed_pattern_kind",
        "legal_mode",
        "access_mode",
        "connector_class",
        "fixture_count",
        "route_level",
        "saved_items",
        "full_gallery_items",
        "count_status",
        "recent_status",
        "varna_status",
        "content_summary",
        "sample_evidence",
    ]
    ws.append(hdr)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in payload["sources"]:
        ws.append(
            [
                row["source_name"],
                row["tier"],
                row["action_lane"],
                row["current_pattern_status"] or "n/a",
                row["proposed_pattern_kind"],
                row["legal_mode"],
                row["access_mode"],
                row["connector_class"],
                row["fixture_count"],
                row["route_evidence"]["route_level"],
                row["saved_items"],
                row["full_gallery_items"],
                row["count_status"] or "",
                row["recent_status"] or "",
                row["varna_status"] or "",
                row["content_summary"],
                row["sample_evidence"],
            ]
        )
        status_cell = ws.cell(ws.max_row, 4)
        if row["current_pattern_status"] == "Patterned":
            status_cell.fill = ok_fill
        elif row["legal_mode"] in {"official_partner_or_vendor_only", "consent_or_manual_only", "legal_review_required", "licensing_required"}:
            status_cell.fill = risk_fill

    ws2 = wb.create_sheet("issues")
    issue_hdr = ["source_name", "issue_id", "title", "detail", "steps"]
    ws2.append(issue_hdr)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in payload["issues"]:
        ws2.append([row[h] for h in issue_hdr])

    ws3 = wb.create_sheet("universality")
    uni_hdr = ["source_name", "status", "saved_items", "services", "categories", "fixture_cases", "reason", "sample_evidence"]
    ws3.append(uni_hdr)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in payload["universality"]:
        ws3.append(
            [
                row["source_name"],
                row["status"],
                row["saved_items"],
                ", ".join(row["services"]),
                ", ".join(row["categories"]),
                ", ".join(row["fixture_cases"]),
                row["reason"],
                row["sample_evidence"],
            ]
        )

    ws4 = wb.create_sheet("candidate_patterns")
    pat_hdr = [
        "source_name",
        "pattern_scope",
        "promotion_gate",
        "proposed_pattern_kind",
        "connector_class",
        "route_level",
        "section_hypotheses",
        "fixture_cases",
        "issues",
    ]
    ws4.append(pat_hdr)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in payload["candidate_patterns"]:
        ws4.append(
            [
                row["source_name"],
                row["pattern_scope"],
                row["promotion_gate"],
                row["proposed_pattern_kind"],
                row["connector_class"],
                row["source_level_pattern"]["route_level"],
                " ; ".join(
                    f"{section['segment_key']} => {section['entry_url'] or 'route missing'} [{section['route_status']}]"
                    for section in row["section_level_patterns"]
                ),
                ", ".join(row["fixture_cases"]),
                " ; ".join(f"{issue['issue_id']} {issue['title']}" for issue in row["issues"]),
            ]
        )

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 80))
            sheet.column_dimensions[col_letter].width = max(14, min(max_len + 2, 60))
    wb.save(OUTPUT_XLSX)


def main() -> None:
    payload = build_registry()
    write_json(payload)
    write_markdown(payload)
    write_candidate_patterns(payload)
    write_xlsx(payload)
    print(f"JSON: {OUTPUT_JSON}")
    print(f"MD: {OUTPUT_MD}")
    print(f"CANDIDATE_JSON: {PATTERN_CANDIDATES_JSON}")
    print(f"CANDIDATE_MD: {PATTERN_CANDIDATES_MD}")
    print(f"XLSX: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
